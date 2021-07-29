import openpyxl as xl
from openpyxl.chart import  BarChart, Reference
from openpyxl.xml.constants import MAX_COLUMN, MAX_ROW, MIN_COLUMN, MIN_ROW
from pathlib import Path

def process_workbook(name):
    wb = xl.load_workbook(name)
    sheet = wb['Sheet1']

    # cell = sheet['a1']
    # cell = sheet.cell(1,1)

    for row in  range(2, sheet.max_row+1):
        cell = sheet.cell(row,3)
        updated_price = cell.value*0.5
        updated_price_cell = sheet.cell(row, 4)
        updated_price_cell.value = updated_price

    values = Reference(sheet, 
                            min_row=2, 
                            max_row=sheet.max_row, 
                            min_col=4, 
                            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'b7')
    wb.save(name)
    


def adding_workbooks():     
    path = Path()
    for file in path.glob('*.xlsx'):
        process_workbook(file)
        print(f'{file} is processed ✔')


adding_workbooks()